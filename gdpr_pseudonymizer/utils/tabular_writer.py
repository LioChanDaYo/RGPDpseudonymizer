"""Structured writers for tabular file formats (Excel, CSV, text).

Writes a TabularDocument back to disk in the specified format,
preserving cell structure from the original file.
"""

from __future__ import annotations

import csv
from pathlib import Path

from gdpr_pseudonymizer.exceptions import FileProcessingError
from gdpr_pseudonymizer.utils.file_handler import HAS_OPENPYXL
from gdpr_pseudonymizer.utils.tabular_reader import TabularDocument


def write_excel(tabular_doc: TabularDocument, output_path: str) -> None:
    """Write a TabularDocument to an Excel (.xlsx) file.

    Creates a new workbook with sheets matching the original structure.
    Formulas are not preserved — output contains static values only.

    Args:
        tabular_doc: Document with cell data to write
        output_path: Path to write the .xlsx file

    Raises:
        FileProcessingError: If openpyxl not installed or write fails
    """
    if not HAS_OPENPYXL:
        raise FileProcessingError(
            "Excel output requires 'openpyxl'. "
            "Install with: pip install gdpr-pseudonymizer[excel]"
        )

    import openpyxl  # type: ignore[import-untyped]

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    try:
        wb = openpyxl.Workbook()
        try:
            # Remove default sheet — we'll create our own
            if wb.worksheets:
                wb.remove(wb.worksheets[0])

            # Create sheets
            sheets_created: dict[str, openpyxl.worksheet.worksheet.Worksheet] = {}
            for sheet_name in tabular_doc.sheet_names:
                ws = wb.create_sheet(title=sheet_name)
                sheets_created[sheet_name] = ws

            # If no sheet names (shouldn't happen), create a default
            if not sheets_created:
                ws = wb.create_sheet(title="Sheet1")
                sheets_created["Sheet1"] = ws

            # Write cell values
            for cell in tabular_doc.cells:
                ws = sheets_created.get(cell.sheet_name)
                if ws is None:
                    ws = wb.create_sheet(title=cell.sheet_name)
                    sheets_created[cell.sheet_name] = ws
                ws.cell(row=cell.row, column=cell.column, value=cell.value)

            wb.save(str(path))
        finally:
            wb.close()
    except FileProcessingError:
        raise
    except Exception as e:
        raise FileProcessingError(f"Cannot write Excel file {output_path}: {e}") from e


def write_csv(tabular_doc: TabularDocument, output_path: str) -> None:
    """Write a TabularDocument to a CSV file.

    Uses comma as delimiter. Multi-sheet documents are flattened
    (all sheets concatenated with an empty row between them).

    Args:
        tabular_doc: Document with cell data to write
        output_path: Path to write the .csv file

    Raises:
        FileProcessingError: If write fails
    """
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    try:
        # Build grid from cells
        rows_by_sheet: dict[str, dict[int, dict[int, str]]] = {}
        for cell in tabular_doc.cells:
            if cell.sheet_name not in rows_by_sheet:
                rows_by_sheet[cell.sheet_name] = {}
            sheet_rows = rows_by_sheet[cell.sheet_name]
            if cell.row not in sheet_rows:
                sheet_rows[cell.row] = {}
            sheet_rows[cell.row][cell.column] = cell.value

        with open(path, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            for sheet_idx, sheet_name in enumerate(tabular_doc.sheet_names):
                if sheet_idx > 0:
                    writer.writerow([])  # separator between sheets
                sheet_rows = rows_by_sheet.get(sheet_name, {})
                if not sheet_rows:
                    continue
                max_row = max(sheet_rows.keys())
                max_col = max(
                    (max(cols.keys()) for cols in sheet_rows.values()), default=0
                )
                for row_num in range(1, max_row + 1):
                    row_data: list[str] = []
                    for col_num in range(1, max_col + 1):
                        row_data.append(sheet_rows.get(row_num, {}).get(col_num, ""))
                    writer.writerow(row_data)
    except PermissionError as e:
        raise FileProcessingError(f"Permission denied: {output_path}") from e
    except OSError as e:
        raise FileProcessingError(f"Cannot write file {output_path}: {e}") from e


def write_tabular_as_text(tabular_doc: TabularDocument, output_path: str) -> None:
    """Write a TabularDocument as flattened plain text.

    Tab-separated cells within rows, newlines between rows,
    double-newlines between sheets.

    Args:
        tabular_doc: Document with cell data to write
        output_path: Path to write the .txt file

    Raises:
        FileProcessingError: If write fails
    """
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    # Build grid from cells
    rows_by_sheet: dict[str, dict[int, dict[int, str]]] = {}
    for cell in tabular_doc.cells:
        if cell.sheet_name not in rows_by_sheet:
            rows_by_sheet[cell.sheet_name] = {}
        sheet_rows = rows_by_sheet[cell.sheet_name]
        if cell.row not in sheet_rows:
            sheet_rows[cell.row] = {}
        sheet_rows[cell.row][cell.column] = cell.value

    sheets_text: list[str] = []
    for sheet_name in tabular_doc.sheet_names:
        sheet_rows = rows_by_sheet.get(sheet_name, {})
        if not sheet_rows:
            continue
        max_row = max(sheet_rows.keys())
        max_col = max((max(cols.keys()) for cols in sheet_rows.values()), default=0)
        rows_text: list[str] = []
        for row_num in range(1, max_row + 1):
            cells: list[str] = []
            for col_num in range(1, max_col + 1):
                cells.append(sheet_rows.get(row_num, {}).get(col_num, ""))
            row_str = "\t".join(cells)
            if row_str.strip():
                rows_text.append(row_str)
        if rows_text:
            sheets_text.append("\n".join(rows_text))

    try:
        with open(path, "w", encoding="utf-8") as f:
            f.write("\n\n".join(sheets_text))
    except PermissionError as e:
        raise FileProcessingError(f"Permission denied: {output_path}") from e
    except OSError as e:
        raise FileProcessingError(f"Cannot write file {output_path}: {e}") from e
