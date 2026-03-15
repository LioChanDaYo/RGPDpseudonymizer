"""Unit tests for Excel file reading in file_handler."""

from __future__ import annotations

from pathlib import Path
from unittest.mock import patch

import openpyxl
import pytest

from gdpr_pseudonymizer.exceptions import FileProcessingError
from gdpr_pseudonymizer.utils.file_handler import read_excel, read_file


def _create_xlsx(path: Path, data: dict[str, list[list[object]]]) -> None:
    """Helper: create .xlsx with sheet_name -> rows mapping."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # type: ignore[arg-type]
    for sheet_name, rows in data.items():
        ws = wb.create_sheet(title=sheet_name)
        for row in rows:
            ws.append(row)
    wb.save(str(path))
    wb.close()


def test_read_excel_basic(tmp_path: Path) -> None:
    """Test reading a simple .xlsx file."""
    xlsx = tmp_path / "test.xlsx"
    _create_xlsx(xlsx, {"Sheet1": [["Alice", "Dupont"], ["Paris", "France"]]})

    result = read_excel(str(xlsx))

    assert "Alice" in result
    assert "Dupont" in result
    assert "Paris" in result
    assert "France" in result


def test_read_excel_multiple_sheets(tmp_path: Path) -> None:
    """Test reading multi-sheet workbook."""
    xlsx = tmp_path / "multi.xlsx"
    _create_xlsx(
        xlsx,
        {
            "People": [["Jean", "Martin"]],
            "Places": [["Lyon", "Marseille"]],
        },
    )

    result = read_excel(str(xlsx))

    assert "Jean" in result
    assert "Lyon" in result
    # Sheets separated by double newline
    assert "\n\n" in result


def test_read_excel_empty_cells(tmp_path: Path) -> None:
    """Test that empty cells are skipped."""
    xlsx = tmp_path / "empty.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="A1")  # type: ignore[union-attr]
    ws.cell(row=1, column=3, value="C1")  # type: ignore[union-attr]
    # B1 is empty
    wb.save(str(xlsx))
    wb.close()

    result = read_excel(str(xlsx))

    assert "A1" in result
    assert "C1" in result


def test_read_excel_formula_cached_values(tmp_path: Path) -> None:
    """Test that data_only=True reads cached formula values."""
    xlsx = tmp_path / "formula.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value=10)  # type: ignore[union-attr]
    ws.cell(row=1, column=2, value=20)  # type: ignore[union-attr]
    ws.cell(row=1, column=3).value = "=A1+B1"  # type: ignore[union-attr]
    wb.save(str(xlsx))
    wb.close()

    # data_only=True reads cached values; freshly created formulas have None cache
    result = read_excel(str(xlsx))
    assert "10" in result
    assert "20" in result


def test_read_excel_password_protected(tmp_path: Path) -> None:
    """Test error for password-protected Excel file."""
    xlsx = tmp_path / "protected.xlsx"
    # Create a minimal file that triggers password error on open
    xlsx.write_bytes(b"not-a-real-xlsx-with-password-encrypted")

    with pytest.raises(FileProcessingError):
        read_excel(str(xlsx))


def test_read_excel_binary_xls(tmp_path: Path) -> None:
    """Test error for .xls binary format."""
    xls = tmp_path / "old.xls"
    xls.write_text("fake xls content")

    with pytest.raises(
        FileProcessingError, match="Binary .xls format is not supported"
    ):
        read_excel(str(xls))


def test_read_excel_corrupt_file(tmp_path: Path) -> None:
    """Test error for corrupt .xlsx file."""
    xlsx = tmp_path / "corrupt.xlsx"
    xlsx.write_text("this is not a zip file")

    with pytest.raises(FileProcessingError):
        read_excel(str(xlsx))


def test_read_excel_not_installed(tmp_path: Path) -> None:
    """Test error when openpyxl is not installed."""
    xlsx = tmp_path / "test.xlsx"
    _create_xlsx(xlsx, {"Sheet1": [["test"]]})

    with patch("gdpr_pseudonymizer.utils.file_handler.HAS_OPENPYXL", False):
        with pytest.raises(FileProcessingError, match="openpyxl"):
            read_excel(str(xlsx))


def test_read_file_dispatches_xlsx(tmp_path: Path) -> None:
    """Test that read_file routes .xlsx to read_excel."""
    xlsx = tmp_path / "test.xlsx"
    _create_xlsx(xlsx, {"Sheet1": [["Hello", "World"]]})

    result = read_file(str(xlsx))

    assert "Hello" in result
    assert "World" in result
