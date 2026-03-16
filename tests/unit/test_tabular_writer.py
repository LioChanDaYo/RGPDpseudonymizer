"""Unit tests for tabular_writer — structured Excel/CSV/text writing."""

from __future__ import annotations

import csv
from pathlib import Path

import openpyxl

from gdpr_pseudonymizer.utils.tabular_reader import CellData, TabularDocument
from gdpr_pseudonymizer.utils.tabular_writer import (
    write_csv,
    write_excel,
    write_tabular_as_text,
)


def _make_doc(
    cells: list[tuple[str, int, int, str, str]],
    sheet_names: list[str] | None = None,
    source_format: str = "xlsx",
) -> TabularDocument:
    """Helper: create a TabularDocument from (sheet, row, col, col_letter, value) tuples."""
    from gdpr_pseudonymizer.utils.tabular_reader import _column_letter

    cell_data = []
    max_row = 0
    max_col = 0
    sheets: list[str] = []
    for sheet, row, col, _col_letter, value in cells:
        col_letter = _column_letter(col)
        cell_ref = f"{sheet}!{col_letter}{row}"
        cell_data.append(
            CellData(
                sheet_name=sheet,
                row=row,
                column=col,
                column_letter=col_letter,
                value=value,
                cell_ref=cell_ref,
            )
        )
        if row > max_row:
            max_row = row
        if col > max_col:
            max_col = col
        if sheet not in sheets:
            sheets.append(sheet)

    return TabularDocument(
        cells=cell_data,
        sheet_names=sheet_names or sheets,
        row_count=max_row,
        column_count=max_col,
        source_format=source_format,
    )


def test_write_excel_basic(tmp_path: Path) -> None:
    """Write TabularDocument to .xlsx, read back and verify."""
    doc = _make_doc(
        [
            ("Sheet1", 1, 1, "A", "Alice"),
            ("Sheet1", 1, 2, "B", "Dupont"),
            ("Sheet1", 2, 1, "A", "Jean"),
            ("Sheet1", 2, 2, "B", "Martin"),
        ]
    )
    output = tmp_path / "output.xlsx"

    write_excel(doc, str(output))

    # Verify
    wb = openpyxl.load_workbook(str(output))
    ws = wb["Sheet1"]
    assert ws.cell(1, 1).value == "Alice"
    assert ws.cell(1, 2).value == "Dupont"
    assert ws.cell(2, 1).value == "Jean"
    assert ws.cell(2, 2).value == "Martin"
    wb.close()


def test_write_csv_basic(tmp_path: Path) -> None:
    """Write TabularDocument to .csv, read back and verify."""
    doc = _make_doc(
        [
            ("Sheet1", 1, 1, "A", "Alice"),
            ("Sheet1", 1, 2, "B", "Dupont"),
            ("Sheet1", 2, 1, "A", "Jean"),
            ("Sheet1", 2, 2, "B", "Martin"),
        ],
        source_format="csv",
    )
    output = tmp_path / "output.csv"

    write_csv(doc, str(output))

    # Verify
    with open(output, encoding="utf-8") as f:
        reader = csv.reader(f)
        rows = list(reader)
    assert rows[0] == ["Alice", "Dupont"]
    assert rows[1] == ["Jean", "Martin"]


def test_write_tabular_as_text(tmp_path: Path) -> None:
    """Verify flattened text output format."""
    doc = _make_doc(
        [
            ("Sheet1", 1, 1, "A", "Alice"),
            ("Sheet1", 1, 2, "B", "Dupont"),
            ("Sheet1", 2, 1, "A", "Jean"),
        ]
    )
    output = tmp_path / "output.txt"

    write_tabular_as_text(doc, str(output))

    text = output.read_text(encoding="utf-8")
    assert "Alice\tDupont" in text
    assert "Jean" in text
